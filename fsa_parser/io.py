from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from .models import LookupResult
from .normalization import normalize_number


OUTPUT_FIELDS = [
    "input_number",
    "registry_type",
    "registry_number",
    "id",
    "url",
    "reg_date",
    "end_date",
    "status_id",
    "status",
    "error",
]

HEADER_MAP = {
    "input_number": "Входной номер",
    "registry_type": "Тип документа",
    "registry_number": "Номер в реестре",
    "id": "Системный ID",
    "url": "Ссылка на карточку",
    "reg_date": "Дата регистрации",
    "end_date": "Дата окончания действия",
    "status_id": "Код статуса",
    "status": "Статус",
    "error": "Ошибка",
}


def read_numbers(path: Path) -> list[str]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv_numbers(path)
    if suffix == ".xlsx":
        return _read_xlsx_numbers(path)
    raise ValueError("Input file must be .csv or .xlsx")


def write_results(path: Path, results: Iterable[LookupResult]) -> None:
    suffix = path.suffix.lower() or ".xlsx"
    if suffix == ".csv":
        _write_csv(path, results)
        return
    if suffix == ".xlsx":
        _write_xlsx(path, results)
        return
    raise ValueError("Output file must be .csv or .xlsx")


def _read_csv_numbers(path: Path) -> list[str]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        sample = file.read(4096)
        file.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample) if sample.strip() else csv.excel
        except csv.Error:
            dialect = csv.excel
        if getattr(dialect, "delimiter", ",") not in {",", ";", "\t", "|"}:
            dialect = csv.excel
        reader = csv.reader(file, dialect)
        rows = list(reader)

    if not rows:
        return []

    header = [normalize_number(cell).lower() for cell in rows[0]]
    has_number_header = "number" in header
    start_index = 1 if has_number_header else 0
    number_index = header.index("number") if has_number_header else 0

    return [
        number
        for row in rows[start_index:]
        if row and (number := normalize_number(row[number_index] if number_index < len(row) else ""))
    ]


def _read_xlsx_numbers(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    workbook.close()

    if not rows:
        return []

    header = [normalize_number(cell).lower() for cell in rows[0]]
    has_number_header = "number" in header
    start_index = 1 if has_number_header else 0
    number_index = header.index("number") if has_number_header else 0

    return [
        number
        for row in rows[start_index:]
        if row and (number := normalize_number(row[number_index] if number_index < len(row) else ""))
    ]


def _write_csv(path: Path, results: Iterable[LookupResult]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        headers = [HEADER_MAP.get(field, field) for field in OUTPUT_FIELDS]
        writer = csv.writer(file)
        writer.writerow(headers)
        for result in results:
            row = result.to_row()
            writer.writerow([row[field] for field in OUTPUT_FIELDS])


def _write_xlsx(path: Path, results: Iterable[LookupResult]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "results"
    
    headers = [HEADER_MAP.get(field, field) for field in OUTPUT_FIELDS]
    worksheet.append(headers)
    
    for result in results:
        row = result.to_row()
        worksheet.append([row[field] for field in OUTPUT_FIELDS])
        
    # Auto-fit columns
    for col in worksheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or "")
            if len(val) > max_len:
                max_len = len(val)
        # Add padding
        worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    workbook.save(path)
